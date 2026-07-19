"""Invite-only guest list (requisition_invites): email normalization, the
CSV/XLSX bulk-upload parser, and the candidate_invite email context. Pure
logic — no DB, no I/O beyond decoding the uploaded bytes."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import datetime

MAX_ROWS = 500
MAX_FILE_BYTES = 1024 * 1024

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Header-row cell → canonical field. Matching is case/space-insensitive.
_HEADER_ALIASES = {
    "email": "email",
    "e-mail": "email",
    "emailaddress": "email",
    "emailid": "email",
    "firstname": "first_name",
    "first": "first_name",
    "givenname": "first_name",
    "lastname": "last_name",
    "last": "last_name",
    "surname": "last_name",
    "familyname": "last_name",
}


def normalize_email(raw: object) -> str | None:
    """Lowercased/trimmed email, or None when it isn't one. This is the
    canonical form stored on requisition_invites and compared at claim."""
    email = str(raw or "").strip().lower()
    return email if _EMAIL_RE.match(email) else None


@dataclass(frozen=True)
class ParsedInvites:
    rows: list[dict]  # {email, first_name, last_name}, emails normalized+deduped
    invalid: list[dict]  # {row: 1-based file row, reason}
    duplicates: int  # in-file duplicate emails skipped


def _rows_from_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        rows: list[list[str]] = []
        for row in wb.active.iter_rows(values_only=True):  # type: ignore[union-attr]
            rows.append(["" if c is None else str(c).strip() for c in row])
            if len(rows) > MAX_ROWS + 1:  # +1 for a header row; cap the scan
                break
        return rows
    finally:
        wb.close()


def _header_map(cells: list[str]) -> dict[str, int] | None:
    """Column map when the first non-empty row is a header row (any cell
    matching an email alias); None → positional email,first,last."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(cells):
        key = re.sub(r"[\s_-]+", "", cell.lower())
        field = _HEADER_ALIASES.get(key)
        if field is not None and field not in mapping:
            mapping[field] = idx
    return mapping if "email" in mapping else None


def parse_invite_file(filename: str, data: bytes) -> ParsedInvites:
    """Parse a bulk-invite upload. Accepts .csv and .xlsx with columns
    email / first name / last name — either a header row (any order, aliased
    names ok) or exactly that positional order without one. Raises ValueError
    with a user-facing message for unusable files."""
    if len(data) > MAX_FILE_BYTES:
        raise ValueError("file too large (max 1 MB)")
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        try:
            raw = _rows_from_xlsx(data)
        except ValueError:
            raise
        except Exception as exc:  # noqa: BLE001 — openpyxl raises a zoo of types
            raise ValueError("could not read the Excel file") from exc
    elif name.endswith(".csv"):
        raw = _rows_from_csv(data)
    else:
        raise ValueError("upload a .csv or .xlsx file")

    first_content = next((i for i, cells in enumerate(raw) if any(cells)), None)
    if first_content is None:
        raise ValueError("the file has no rows")
    columns = _header_map(raw[first_content])
    start = first_content + 1 if columns is not None else first_content
    if columns is None:
        columns = {"email": 0, "first_name": 1, "last_name": 2}

    def cell(cells: list[str], field: str) -> str:
        idx = columns.get(field)
        return cells[idx] if idx is not None and idx < len(cells) else ""

    rows: list[dict] = []
    invalid: list[dict] = []
    duplicates = 0
    seen: set[str] = set()
    for i, cells in enumerate(raw[start:], start=start + 1):
        if not any(cells):
            continue
        email = normalize_email(cell(cells, "email"))
        if email is None:
            invalid.append({"row": i, "reason": "invalid email"})
            continue
        if email in seen:
            duplicates += 1
            continue
        seen.add(email)
        rows.append(
            {
                "email": email,
                "first_name": cell(cells, "first_name"),
                "last_name": cell(cells, "last_name"),
            }
        )
        if len(rows) > MAX_ROWS:
            raise ValueError(f"too many rows (max {MAX_ROWS} invites per upload)")
    return ParsedInvites(rows=rows, invalid=invalid, duplicates=duplicates)


def candidate_invite_context(
    *,
    org_name: str,
    interview_name: str,
    interview_url: str,
    first_name: str = "",
    closes_at: datetime | None = None,
) -> dict:
    """Context for the candidate_invite template. valid_until/candidate_name
    render conditionally, so empty strings simply drop those lines."""
    return {
        "org_name": org_name,
        "interview_name": interview_name,
        "interview_url": interview_url,
        "candidate_name": first_name.strip(),
        "valid_until": f"{closes_at:%B} {closes_at.day}, {closes_at.year}" if closes_at else "",
    }
